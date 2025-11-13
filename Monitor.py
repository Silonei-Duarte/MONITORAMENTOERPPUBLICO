import sys
import os
import threading
import configparser
import time
import glob
import oracledb
import openpyxl
from datetime import datetime
from PyQt6.QtWidgets import QApplication, QLabel, QMainWindow, QTreeView, QPushButton, QMessageBox, QHeaderView, \
    QAbstractItemView
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QIcon, QStandardItemModel, QStandardItem, QFont, QPalette, QColor
from Dashboard import Dashboards
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from email.utils import make_msgid
import cryptography
from cryptography.fernet import Fernet
import shutil
from Configuracao import Configuracao
from Legenda import dados_legenda, Legendas
from Sobre import Sobre

processos_lista = []
licencas = []
class AplicativoMonitoramento(QMainWindow):
    SQL_QUERY_CONEXOES = """
        SELECT R911MOD.NumSec AS NUMSEC,
               R911SEC.APPUSR AS APPUSR, 
               LISTAGG(R911MOD.ModNam, ', ') WITHIN GROUP (ORDER BY R911MOD.ModNam) AS ModNam
        FROM R911MOD
        JOIN R911SEC ON R911MOD.NumSec = R911SEC.NumSec
        GROUP BY R911MOD.NumSec, R911SEC.APPUSR
        ORDER BY R911SEC.APPUSR, R911MOD.NumSec
    """

    def __init__(self):
        super().__init__()
        self.setWindowTitle("Monitor de Uso de Processos")
        self.setGeometry(600, 50, 670, 950)
        self.setWindowIcon(QIcon(os.path.join(os.path.dirname(__file__), 'icone.ico')))

        # Define o caminho para o arquivo db.ini
        self.caminho_arquivo_ini = os.path.join(os.path.dirname(__file__), 'db.ini')

        # Obter apims de criptografia e inicializar Fernet
        self.key = self.obter_apims_criptografia()
        self.fernet = Fernet(self.key)

        # Carregar a configuração, incluindo processos_lista e licencas
        self.usuarios_permitidos, self.horarios_desconexoes, self.minutos, self.horaemail, self.email, self.destino, self.legenda = self.carregar_configuracao()

        # Botões
        self.btn_exibir_conexoes = QPushButton("Exibir Conexões", self)
        self.btn_exibir_conexoes.setGeometry(530, 20, 120, 35)
        self.btn_exibir_conexoes.setStyleSheet("background-color: #3e3e3e; color: #ffffff;")
        self.btn_exibir_conexoes.clicked.connect(self.exibir_conexoes)

        self.btn_derrubar_conexao = QPushButton("Derrubar Conexão", self)
        self.btn_derrubar_conexao.setGeometry(530, 70, 120, 35)
        self.btn_derrubar_conexao.setStyleSheet("background-color: #3e3e3e; color: #ffffff;")
        self.btn_derrubar_conexao.clicked.connect(self.derrubar_conexao)

        self.btn_abrir_dashboard = QPushButton("Dashboard", self)
        self.btn_abrir_dashboard.setGeometry(530, 120, 120, 35)
        self.btn_abrir_dashboard.setStyleSheet("background-color: #3e3e3e; color: #ffffff;")
        self.btn_abrir_dashboard.clicked.connect(self.abrir_dashboard)

        self.btn_abrir_legenda = QPushButton("Legendas", self)
        self.btn_abrir_legenda.setGeometry(530, 170, 120, 35)
        self.btn_abrir_legenda.setStyleSheet("background-color: #3e3e3e; color: #ffffff;")
        self.btn_abrir_legenda.clicked.connect(self.abrir_legenda)

        self.btn_abrir_sobre = QPushButton("Sobre", self)
        self.btn_abrir_sobre.setGeometry(530, 880, 120, 35)
        self.btn_abrir_sobre.setStyleSheet("background-color: #3e3e3e; color: #ffffff;")
        self.btn_abrir_sobre.clicked.connect(self.abrir_sobre)


        self.btn_abrir_configuracao = QPushButton("Configurações", self)
        self.btn_abrir_configuracao.setGeometry(530, 220, 120, 35)
        self.btn_abrir_configuracao.setStyleSheet("background-color: #3e3e3e; color: #ffffff;")
        self.btn_abrir_configuracao.clicked.connect(self.abrir_configuracao)

        # Caixa para exibir a contagem de processos
        self.label_contagem_processos = QLabel(self)
        self.label_contagem_processos.setGeometry(530, 270, 120, 450)
        self.label_contagem_processos.setWordWrap(True)
        self.label_contagem_processos.setStyleSheet("color: #ffffff;")

        # Configura a paleta de cores do aplicativo para o tema escuro
        palette = QPalette()
        palette.setColor(QPalette.ColorRole.Window, QColor("#2e2e2e"))
        palette.setColor(QPalette.ColorRole.Button, QColor("#3e3e3e"))
        palette.setColor(QPalette.ColorRole.ButtonText, QColor("#ffffff"))
        self.setPalette(palette)

        # Configuração da caixa de Conexoes
        self.tree = QTreeView(self)
        self.tree.setGeometry(20, 20, 500, 900)
        self.tree.setHeaderHidden(False)
        self.tree.setFont(QFont("Segoe UI", 10))

        # Model para o Caixa de Conexoes
        self.model = QStandardItemModel()
        self.model.setHorizontalHeaderLabels(["      Conexão", "Usuário", "Processos"])
        self.tree.setModel(self.model)
        self.tree.setSelectionBehavior(QTreeView.SelectionBehavior.SelectRows)
        self.tree.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)

        # Ajuste do cabeçalho e largura das colunas
        self.aplicar_estilo_colconex()

        # Define o estilo para o QTreeView
        self.tree.setStyleSheet(
            "QTreeView { background-color: #2e2e2e; color: #ffffff; } "
            "QTreeView::item:selected { background-color: #FF4500; color: #ffffff; } "
            "QHeaderView::section { background-color: #1e1e1e; color: #ffffff; } "
            "QTreeView::item:hover { background-color: #FFA500; color: #ffffff; }"
        )

        # Ajuste da barra de rolagem
        self.tree.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.tree.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)

        # Inicia a thread para atualização periódica
        self.intervalo = 60
        self.thread = threading.Thread(target=self.atualizar_dados_periodicamente, args=(self.intervalo,))
        self.thread.daemon = True
        self.thread.start()

    def aplicar_estilo_colconex(self):
        header = self.tree.header()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Fixed)  # Define o modo de redimensionamento fixo
        header.resizeSection(0, 80)  # Largura da coluna "Conexão"
        header.resizeSection(1, 130)  # Largura da coluna "Usuário"
        header.resizeSection(2, 250)  # Largura da coluna "Processos"

        # Centraliza o texto nas células do TreeView
        for row in range(self.model.rowCount()):
            for column in range(self.model.columnCount()):
                item = self.model.item(row, column)
                if item:
                    item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

        # Centraliza o texto do cabeçalho
        header.setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)

    def abrir_dashboard(self):
        try:
            self.dashboard = Dashboards()  # Cria uma instância da classe Dashboard
            self.dashboard.show()  # Exibe a janela
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao abrir o Dashboard: {e}")

    def abrir_legenda(self):
        try:
            self.legenda = Legendas()
            self.legenda.show()
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao abrir a Legenda {e}")

    def abrir_sobre(self):
        try:
            self.sobre = Sobre()
            self.sobre.show()
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao abrir Sobre {e}")

    def abrir_configuracao(self):
        try:
            self.configuracao = Configuracao()
            self.configuracao.show()
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao abrir Configurações {e}")

    def gerar_nome_arquivo(self):
        return f"Senior_{datetime.now().strftime('%d.%m.%y')}.xlsx"

    def obter_apims_criptografia(self):
        """Carrega a apims de criptografia existente."""
        apims_arquivo = os.path.join(os.path.dirname(__file__), 'apims.key')

        if not os.path.exists(apims_arquivo):
            return None

        try:
            with open(apims_arquivo, 'rb') as f:
                apims = f.read()
            return apims
        except Exception:
            return None

    def conectar_bd(self):
        """Estabelece a conexão com o banco de dados Oracle."""
        config = configparser.ConfigParser()

        if not os.path.exists(self.caminho_arquivo_ini):
            return None

        config.read(self.caminho_arquivo_ini)

        try:
            # Acessa as credenciais criptografadas e descriptografa
            ip_criptografado = config.get('CONEXAO', 'ip')
            porta_criptografado = config.get('CONEXAO', 'porta')
            service_name_criptografado = config.get('CONEXAO', 'service_name')
            usuario_criptografado = config.get('CONEXAO', 'usuario')
            senha_criptografada = config.get('CONEXAO', 'senha')

            # Descriptografa os dados
            ip = self.fernet.decrypt(ip_criptografado.encode()).decode()
            porta = self.fernet.decrypt(porta_criptografado.encode()).decode()
            service_name = self.fernet.decrypt(service_name_criptografado.encode()).decode()
            usuario = self.fernet.decrypt(usuario_criptografado.encode()).decode()
            senha = self.fernet.decrypt(senha_criptografada.encode()).decode()

            # Conecta ao banco de dados com o service name configurado
            dsn_tns = oracledb.makedsn(ip, porta, service_name=service_name)
            connection = oracledb.connect(user=usuario, password=senha, dsn=dsn_tns)

            return connection
        except Exception:
            return None

    def carregar_configuracao(self):
        """Lê o arquivo db.ini e carrega a lista de processos, licenças e outras configurações."""
        config = configparser.ConfigParser()

        if not os.path.exists(self.caminho_arquivo_ini):
            return None

        config.read(self.caminho_arquivo_ini)

        try:
            global processos_lista, licencas

            # Carrega processos e licenças
            processos_str = config.get('PROPRIETARIA', 'processos', fallback='')
            licencas_str = config.get('PROPRIETARIA', 'licencas', fallback='')

            processos_lista = [proc.strip() for proc in processos_str.split(',') if proc.strip()]
            licencas = [lic.strip() for lic in licencas_str.split(',') if lic.strip()]

            if "SEG. TELA" not in processos_lista:
                processos_lista.append("SEG. TELA")

            # Carrega a legenda
            legenda_str = config.get('LEGENDA', 'legenda', fallback='')
            legenda = []
            if legenda_str:
                linhas_legenda = [linha.strip() for linha in legenda_str.split('\n') if linha.strip()]
                for linha in linhas_legenda:
                    if ',' in linha:
                        sigla, descricao = linha.split(',', 1)
                        legenda.append(
                            (sigla.strip(), descricao.strip().rstrip(',')))

            # Carrega outras configurações
            usuarios_str = config.get('PERMITIDOS', 'Usuarios', fallback='')
            usuarios = [usuario.strip() for usuario in usuarios_str.split(',') if usuario.strip()]

            horarios_str = config.get('HORARIOS DESCONEXOES', 'Horarios', fallback='')
            horarios = [horario.strip() for horario in horarios_str.split(',') if horario.strip()]

            minutos_str = config.get('MINUTOS PARA DESCONEXAO', 'Minutos', fallback='9999')
            minutos = int(minutos_str) if minutos_str.isdigit() else 9999

            horaemail_str = config.get('HORARIOS E-MAILS', 'Horarios', fallback='')
            horaemail = [hora.strip() for hora in horaemail_str.split(',') if hora.strip()]

            email_str = config.get('EMAILS', 'E-mails', fallback='')
            email = [email.strip() for email in email_str.split(';') if email.strip()]

            destino = config.get('DESTINO', 'Destino', fallback='').strip()

            return usuarios, horarios, minutos, horaemail, email, destino, legenda

        except configparser.NoSectionError:
            return None

    def carregar_planilha(self):
        """Carrega os dados da planilha para continuar o controle do tempo acumulado."""
        nome_arquivo = self.gerar_nome_arquivo()
        if os.path.exists(nome_arquivo):
            wb = openpyxl.load_workbook(nome_arquivo)
            ws = wb.active
            tempo_acumulado = {}

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                user = row[0].value
                tempos = [cell.value for cell in row[1:]]
                if user:
                    tempo_acumulado[user] = {}
                    for i, proc in enumerate(processos_lista):
                        tempo_str = str(tempos[i])  # Converte o valor para string
                        if tempo_str:
                            try:
                                # Converte o tempo para segundos
                                horas, minutos, segundos = map(int, tempo_str.split(':'))
                                tempo_acumulado[user][proc] = horas * 3600 + minutos * 60 + segundos
                            except ValueError:
                                # Se a conversão falhar, considera o tempo como 0
                                tempo_acumulado[user][proc] = 0
                        else:
                            tempo_acumulado[user][proc] = 0

            return tempo_acumulado
        return {}

    def formatar_tempo(self, segundos):
        """Formata o tempo em horas, minutos e segundos no formato HH:MM:SS."""
        horas, segundos = divmod(segundos, 3600)
        minutos, segundos = divmod(segundos, 60)
        return f"{horas:02}:{minutos:02}:{segundos:02}"

    def atualizar_planilha(self, dados, intervalo):
        """Atualiza a planilha com os dados coletados, adicionando o tempo de uso e o tempo de 'Segurando Tela'."""
        # Define o caminho do arquivo .ini no mesmo diretório do script
        self.caminho_arquivo_ini = os.path.join(os.path.dirname(__file__), 'db.ini')
        # Carrega os usuários permitidos e horários de desconexão uma vez
        self.usuarios_permitidos, self.horarios_desconexoes, self.minutos, self.horaemail, self.email, self.destino, self.legenda = self.carregar_configuracao()

        global processos, appusr
        nome_arquivo = self.gerar_nome_arquivo()
        tempo_acumulado = self.carregar_planilha()

        try:
            wb = openpyxl.load_workbook(nome_arquivo)
        except FileNotFoundError:
            wb = openpyxl.Workbook()

        ws = wb.active

        # Envio de E-mails nos  Horarios Agendado
        hora_atual = time.strftime("%H:%M")
        if hora_atual in self.horaemail:
            self.executar_sql_e_enviar_email()

        # Desconexão nos Horarios Agendados
        hora_atual = time.strftime("%H:%M")
        if hora_atual in self.horarios_desconexoes:
            self.desconexao_agendada()

        # Verifica se o cabeçalho está presente na primeira linha
        if ws.cell(row=1, column=1).value != "Usuário" or ws.cell(row=1, column=2).value != "All":
            # Se o cabeçalho não estiver correto, mova as linhas existentes para baixo
            if ws.max_row > 1:
                ws.insert_rows(1)
            # Adiciona o cabeçalho
            ws.cell(row=1, column=1, value="Usuário")
            for col, proc in enumerate(processos_lista, start=2):
                ws.cell(row=1, column=col, value=proc)
            ws.cell(row=1, column=len(processos_lista) + 2, value="ULTS.PROC.")
            ws.cell(row=1, column=len(processos_lista) + 3, value="PEN.PROC.")
            ws.cell(row=1, column=len(processos_lista) + 4, value="CONTADOR")

        usuarios_processos = {}
        usuarios_presentes = set()

        for numsec, appusr, ModNam in dados:
            usuarios_presentes.add(appusr)
            processos = [proc.strip() for proc in ModNam.split(',')]  # Divide a string em processos individuais
            if appusr not in usuarios_processos:
                usuarios_processos[appusr] = {}
            for proc in processos:
                if proc not in usuarios_processos[appusr]:
                    usuarios_processos[appusr][proc] = []
                usuarios_processos[appusr][proc].append(numsec)

        for appusr, processos in usuarios_processos.items():
            # Garantir que o usuário tenha uma entrada em tempo_acumulado
            if appusr not in tempo_acumulado:
                tempo_acumulado[appusr] = {}

            for proc, numsecs in processos.items():
                # Inicializar o tempo acumulado do processo se não existir
                if proc not in tempo_acumulado[appusr]:
                    tempo_acumulado[appusr][proc] = 0
                # Adicionar o intervalo ao processo específico
                tempo_acumulado[appusr][proc] += intervalo

            # Atualiza o tempo de 'Segurando Tela'
            self.atualizar_seg_tela(tempo_acumulado, appusr, processos, intervalo)

            linha_encontrada = False
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if row[0].value == appusr:
                    linha_encontrada = True
                    # Atualizar o tempo dos processos na planilha
                    for i, proc in enumerate(processos_lista):
                        tempo_atual = tempo_acumulado.get(appusr, {}).get(proc, 0)
                        ws.cell(row=row[0].row, column=i + 2, value=self.formatar_tempo(tempo_atual))

            if not linha_encontrada:
                # Adiciona nova linha com tempo acumulado
                row = [appusr] + [self.formatar_tempo(tempo_acumulado.get(appusr, {}).get(proc, 0)) for proc in
                                  processos_lista]
                ult_proc = ', '.join(processos.keys())
                row.append(ult_proc)
                ws.append(row)
            self.atualizar_ults_proc(ws, appusr, processos)

        # Limpando as células de ULTS.PROC. para usuários não presentes nesta execução
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if row[0].value not in usuarios_presentes:
                ws.cell(row=row[0].row, column=len(processos_lista) + 2, value=" ")

        # Atualizar PEN.PROC. e CONTADOR E DERRUBA COM BASE NO TEMPO DE CONEXAO
        self.atualizar_contador(ws)
        self.ler_planilha_e_filtrar_usuarios(ws)

        wb.save(nome_arquivo)
        self.copiar_arquivo(self.destino)

    def ler_planilha_e_filtrar_usuarios(self, ws):
        """Filtra usuários com contador maior ou igual a 60 e executa o comando SQL de Derrubar para esses usuários, exceto se estiverem na lista de usuários permitidos."""
        contador_col = len(processos_lista) + 4

        # Coleta os usuários com contador maior ou igual a 60
        usuarios_filtrados = [row[0] for row in ws.iter_rows(min_row=2, max_col=ws.max_column, values_only=True)
                              if row[contador_col - 1] >= self.minutos]

        # Filtra usuários permitidos
        usuarios_permitidos = set(self.usuarios_permitidos)
        usuarios_a_excluir = [usuario for usuario in usuarios_filtrados if usuario not in usuarios_permitidos]

        if not usuarios_a_excluir:
            return

        # Constrói a cláusula IN para o comando SQL
        usuarios_lista = "', '".join(usuarios_a_excluir)
        comando_sql = f"""
        DELETE FROM R911MOD
        WHERE NUMSEC IN (
            SELECT R911MOD.NUMSEC
            FROM R911MOD
            JOIN R911SEC ON R911MOD.NumSec = R911SEC.NumSec
            WHERE R911SEC.APPUSR IN ('{usuarios_lista}')
        )
        """

        # Executa o comando SQL
        try:
            with self.conectar_bd() as conn:
                with conn.cursor() as cursor:
                    cursor.execute(comando_sql)
                    conn.commit()
        except oracledb.DatabaseError as e:
            # Trate o erro de banco de dados se necessário
            pass

    def atualizar_seg_tela(self, tempo_acumulado, appusr, processos, intervalo):
        """Atualiza o tempo de 'Segurando Tela' para o usuário específico."""
        # Criar um dicionário para armazenar as conexões por processo
        processos_atuais = {}

        for proc in processos.keys():
            if proc not in processos_atuais:
                processos_atuais[proc] = set()
            for numsec in processos[proc]:
                processos_atuais[proc].add(numsec)

        # Verificar se o usuário tem o mesmo processo em diferentes conexões
        for proc, conexoes in processos_atuais.items():
            if len(conexoes) > 1:
                if "SEG. TELA" not in tempo_acumulado[appusr]:
                    tempo_acumulado[appusr]["SEG. TELA"] = 0
                tempo_acumulado[appusr]["SEG. TELA"] += intervalo
            else:
                # não zera o acumulado, apenas ignora se não está segurando tela
                if "SEG. TELA" not in tempo_acumulado[appusr]:
                    tempo_acumulado[appusr]["SEG. TELA"] = 0

    def atualizar_ults_proc(self, ws, appusr, processos):
        """Atualiza a coluna ULTS.PROC. para o usuário específico, salvando todos os processos separados por vírgulas ou colocando '-' se não houver processos abertos."""
        # Converte os processos em uma string separada por vírgulas
        ult_proc = ', '.join(processos.keys()) if processos else "-"

        # Percorre as linhas da planilha para encontrar o usuário
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if row[0].value == appusr:
                # Atualiza a célula com os processos na coluna ULTS.PROC.
                ws.cell(row=row[0].row, column=len(processos_lista) + 2, value=ult_proc)
                break

    def atualizar_contador(self, ws):
        """Imprime o conteúdo das colunas ULTS.PROC. e PEN.PROC. para cada usuário/linha, verifica se são iguais ou diferentes, e atualiza a coluna CONTADOR e PEN.PROC."""
        ult_proc_col = len(processos_lista) + 2  # Coluna "ULTS.PROC."
        pen_proc_col = len(processos_lista) + 3  # Coluna "PEN.PROC."
        contador_col = len(processos_lista) + 4  # Coluna "CONTADOR"

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            usuario = row[0].value
            ult_proc = row[ult_proc_col - 1].value  # Índice baseado em 1
            pen_proc = row[pen_proc_col - 1].value  # Índice baseado em 1

            if usuario:
                # Verifica se ULTS.PROC. e PEN.PROC. estão vazios
                if ult_proc == " " and pen_proc == " ":
                    ws.cell(row=row[0].row, column=contador_col, value=0)
                elif ult_proc == pen_proc:
                    # Incrementa o valor da coluna CONTADOR
                    contador_valor = row[contador_col - 1].value or 0
                    ws.cell(row=row[0].row, column=contador_col, value=contador_valor + 1)
                    # Copia o conteúdo de ULTS.PROC. para PEN.PROC.
                    ws.cell(row=row[0].row, column=pen_proc_col, value=ult_proc)
                else:
                    # Define o valor da coluna CONTADOR como 0
                    ws.cell(row=row[0].row, column=contador_col, value=0)
                    # Atualiza a coluna PEN.PROC. com o valor de ULTS.PROC.
                    ws.cell(row=row[0].row, column=pen_proc_col, value=ult_proc)

    def copiar_arquivo(self, destino):
        """Copia o arquivo Senior_dd.mm.yy.xlsx para o destino especificado e mantém apenas as últimas 7 planilhas na origem e no destino."""
        # Gera o nome do arquivo com base na data atual
        nome_arquivo = f"Senior_{datetime.now().strftime('%d.%m.%y')}.xlsx"
        caminho_origem = os.path.join(os.getcwd(), nome_arquivo)
        caminho_destino = os.path.join(destino, nome_arquivo)

        try:
            shutil.copy(caminho_origem, caminho_destino)
        except Exception:
            pass  # Ignora qualquer erro

        # Chama a função para limpar os diretórios
        self.manter_ultimas_planilhas(os.getcwd())
        self.manter_ultimas_planilhas(destino)

    def manter_ultimas_planilhas(self, caminho_diretorio, max_arquivos=7):
        """Mantém apenas os últimos `max_arquivos` arquivos no diretório especificado, excluindo os mais antigos."""
        padrao_arquivos = os.path.join(caminho_diretorio, "Senior_*.xlsx")
        arquivos = glob.glob(padrao_arquivos)

        # Ordena por data de modificação e, em caso de empate, por nome de arquivo
        arquivos.sort(key=lambda x: (os.path.getctime(x), x), reverse=True)

        arquivos_a_excluir = arquivos[max_arquivos:]

        for arquivo in arquivos_a_excluir:
            try:
                os.remove(arquivo)
            except Exception:
                pass  # Ignora erros ao excluir arquivos
    def atualizar_dados_periodicamente(self, intervalo):
        proximo_inicio = time.perf_counter()  # Define o tempo para o próximo loop
        while True:
            inicio_loop = time.perf_counter()

            connection = self.conectar_bd()
            if connection is None:
                time.sleep(intervalo)
                continue

            try:
                cursor = connection.cursor()
                cursor.execute(self.SQL_QUERY_CONEXOES)
                dados = cursor.fetchall()
                self.atualizar_planilha(dados, intervalo)
            except Exception:
                pass  # Ignora qualquer erro de coleta de dados
            finally:
                cursor.close()
                connection.close()

            # Calcula o tempo que o loop levou para ser executado
            duracao_execucao = time.perf_counter() - inicio_loop

            # Define o tempo do próximo início baseado no intervalo fixo
            proximo_inicio += intervalo

            # Calcula o tempo até o próximo ciclo
            tempo_restante = proximo_inicio - time.perf_counter()

            # Se o tempo restante for maior que zero, aguarda o tempo exato
            if tempo_restante > 0:
                time.sleep(tempo_restante)
            else:
                # Caso o loop tenha demorado mais que o intervalo, ajusta o próximo ciclo
                proximo_inicio = time.perf_counter()  # Reajusta o próximo início

    def exibir_conexoes(self):
        connection = self.conectar_bd()
        if connection is None:
            return

        try:
            cursor = connection.cursor()
            cursor.execute(self.SQL_QUERY_CONEXOES)

            # Limpar o TreeView
            self.model.clear()
            self.model.setHorizontalHeaderLabels(["      Conexão", "Usuário", "Processos"])

            # Dicionário para armazenar processos por usuário
            usuarios_processos = {}

            for row in cursor.fetchall():
                numsec, appusr, ModNam = row
                item_numsec = QStandardItem(str(numsec))
                item_user = QStandardItem(appusr)
                item_processes = QStandardItem(ModNam)
                self.model.appendRow([item_numsec, item_user, item_processes])

                # Armazenar processos do usuário
                if appusr not in usuarios_processos:
                    usuarios_processos[appusr] = []
                usuarios_processos[appusr].append(ModNam)

            # Atualizar a contagem de processos na label
            self.exibir_contagem_processos(usuarios_processos)

        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao exibir conexões: {e}")
        finally:
            self.aplicar_estilo_colconex()
            cursor.close()
            connection.close()

    def exibir_contagem_processos(self, usuarios_processos):
        # Primeira consulta SQL
        SQL_QUERY_LICENCAS = """
        SELECT COUNT(DISTINCT R911SEC.APPUSR) AS DistinctAPPUSRCount
        FROM R911MOD
        JOIN R911SEC ON R911MOD.NumSec = R911SEC.NumSec
        """

        # Segunda consulta SQL para "SEG. TELA"
        SQL_QUERY_SEG_TELA = """
        SELECT COUNT(DISTINCT APPUSR) AS APPUSRCount
        FROM (
        SELECT R911SEC.APPUSR
        FROM R911MOD
        JOIN R911SEC ON R911MOD.NumSec = R911SEC.NumSec
        GROUP BY R911SEC.APPUSR, R911MOD.ModNam
        HAVING COUNT(DISTINCT R911SEC.NumSec) > 1
)
        """

        contagem_usuarios = None
        contagem_seg_tela = None

        conexao = self.conectar_bd()  # Estabelece a conexão com o banco de dados

        if conexao:
            try:
                with conexao.cursor() as cursor:
                    # Executa a primeira consulta
                    cursor.execute(SQL_QUERY_LICENCAS)
                    resultado = cursor.fetchone()
                    contagem_usuarios = resultado[0] if resultado else None

                    # Executa a segunda consulta apenas se o processo for "SEG. TELA"
                    cursor.execute(SQL_QUERY_SEG_TELA)
                    resultado_seg_tela = cursor.fetchone()
                    contagem_seg_tela = resultado_seg_tela[0] if resultado_seg_tela else None
            except Exception:
                pass
            finally:
                conexao.close()  # Fecha a conexão

        licencas.append(f"{contagem_usuarios}")  # Inclui o valor da primeira consulta na lista

        # Inicia a construção do texto de contagem das licenças
        contagem_texto = "Licenças: \n\n"
        for i, processo in enumerate(processos_lista):
            if processo == "SEG. TELA" and contagem_seg_tela is not None:
                contagem_texto += f"{processo}: {contagem_seg_tela} / {licencas[i]}\n"
            else:
                contagem = sum(
                    processo in modnames for processos in usuarios_processos.values() for modnames in processos)
                contagem_texto += f"{processo}: {contagem} / {licencas[i]}\n"

        self.label_contagem_processos.setText(contagem_texto)

    def derrubar_conexao(self):
        selected_indexes = self.tree.selectedIndexes()
        if not selected_indexes:
            QMessageBox.warning(self, "Seleção Inválida", "Selecione uma linha para derrubar a conexão.")
            return

        numsec = selected_indexes[0].data()
        connection = self.conectar_bd()
        if connection is None:
            return

        try:
            cursor = connection.cursor()
            cursor.execute("DELETE FROM R911MOD WHERE numsec = :numsec", {'numsec': numsec})
            connection.commit()
            QMessageBox.information(self, "Sucesso", f"Conexão {numsec} derrubada com sucesso.")
            self.exibir_conexoes()  # Atualiza a lista de conexões
        finally:
            cursor.close()
            connection.close()

    def carregar_configuracao_smtp(self):
        """Carrega as configurações do servidor SMTP a partir do arquivo db.ini"""
        config = configparser.ConfigParser()
        config.read(self.caminho_arquivo_ini)

        # Verifica se a seção SMTP existe e carrega os valores
        if 'SMTP' in config:
            self.smtp_server = config.get('SMTP', 'smtp_server')
            self.smtp_port = config.get('SMTP', 'smtp_port')
            self.smtp_user = config.get('SMTP', 'smtp_user')

            # Descriptografa a senha do SMTP
            smtp_password_criptografada = config.get('SMTP', 'smtp_password')
            self.smtp_password = self.fernet.decrypt(smtp_password_criptografada.encode()).decode()

    def executar_sql_e_enviar_email(self):
        """Executa a consulta SQL, formata os resultados como HTML e envia por e-mail."""
        comando_sql = self.SQL_QUERY_CONEXOES

        # Executa a consulta SQL e coleta os resultados
        resultados_html = ""
        try:
            conn = self.conectar_bd()
            if conn:
                try:
                    with conn.cursor() as cursor:
                        cursor.execute(comando_sql)
                        resultados = cursor.fetchall()

                        # Renomeia as colunas para "Conexão", "Usuário" e "Processos"
                        colunas = ["Conexão", "Usuário", "Processos"]

                    # Formata os resultados como tabela HTML
                    resultados_tabela = "<table border='1' style='border-collapse: collapse; width: auto;'>"
                    resultados_tabela += "<tr>" + "".join(f"<th>{coluna}</th>" for coluna in colunas) + "</tr>"
                    for row in resultados:
                        resultados_tabela += "<tr>" + "".join(f"<td>{cell}</td>" for cell in row) + "</tr>"
                    resultados_tabela += "</table>"
                finally:
                    conn.close()
        except Exception:
            pass

        # Adiciona a legenda ao HTML usando a variável carregada
        legenda_html = "<table border='1' style='border-collapse: collapse; width: auto;'>"
        legenda_html += "<tr><th>Sigla</th><th>Descrição</th></tr>"
        for sigla, descricao in self.legenda:
            legenda_html += f"<tr><td>{sigla}</td><td>{descricao}</td></tr>"
        legenda_html += "</table>"

        # Nome do título igual ao assunto
        titulo_email = f"Usuários Conectados às {datetime.now().strftime('%H:%M')} no Senior."

        # Formatação HTML com a tabela principal ajustada para ocupar apenas o espaço necessário
        # Formatação HTML com a tabela principal ajustada para ocupar apenas o espaço necessário
        resultados_html = f"""
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; font-size: 13px; color: #333333; margin: 0; padding: 0; }}
                .container {{ display: inline-block; padding: 10px; max-width: auto; border: 1px solid #dddddd; }}
                table {{ border-collapse: collapse; margin: 0; }}
                th, td {{ border: 1px solid #777777; padding: 4px; text-align: left; font-size: 12px; color: #333333; white-space: nowrap; }}
                th {{ background-color: #666666; color: #ffffff; }}
                a {{ color: #2196f3; font-weight: bold; text-decoration: none; font-size: 13px; }} /* Azul mais fraco para o link e fonte 13px */
                .yellow-text {{ color: yellow; font-weight: bold; font-size: 13px; }} /* Cor amarela para o texto de aviso, tamanho 13px */
                .bold-text {{ font-weight: bold; font-size: 13px; }} /* Texto em negrito e tamanho 13px */
            </style>
        </head>
        <body>
            <table style='border-collapse: collapse; margin: auto;'>
                <tr>
                    <td colspan="2" style="text-align: center; padding: 10px; background-color: #444444; color: #ffffff;">
                        <h1 style="font-size: 18px; margin: 0;">{titulo_email}</h1>
                    </td>
                </tr>
                <tr>
                    <td style="padding-right: 20px; vertical-align: top;">
                        {resultados_tabela}
                    </td>
                    <td style="vertical-align: top;">
                        {legenda_html}
                    </td>
                </tr>
                <tr>
                    <td colspan="2" style="text-align: center; padding: 10px; background-color: #444444; color: #cccccc;">
                        <span class="bold-text">Acesse o relatório completo com filtros de uso do sistema dos últimos 7 dias utilizando o programa Dashboard abaixo:</span><br>
                        <a href="{self.destino}">{self.destino}</a><br><br>
                        <span class="yellow-text">Aguarde de 15 segundos a 2 minutos para o carregamento do programa.</span><br>
                        <span class="yellow-text">Ajuste a escala de resolução do seu computador de 125% para 100% se necessário.</span><br><br>
                        <span style="font-size: 12px;">Este é um e-mail automático. Para mais informações, entre em contato com a equipe de suporte.</span>
                    </td>
                </tr>
            </table>
        </body>
        </html>
        """


        # Carrega as configurações do servidor SMTP
        self.carregar_configuracao_smtp()

        # Envia o e-mail com os resultados usando SMTP
        try:
            # Configurações do e-mail
            destinatario = "; ".join(self.email)
            assunto = titulo_email

            # Cria a mensagem do e-mail
            msg = MIMEMultipart()
            msg['From'] = self.smtp_user
            msg['To'] = destinatario
            msg['Subject'] = assunto

            # Adiciona o corpo do e-mail em HTML
            msg.attach(MIMEText(resultados_html, 'html'))

            # Envia o e-mail via servidor SMTP com SSL/TLS
            with smtplib.SMTP_SSL(self.smtp_server, int(self.smtp_port)) as server:
                server.login(self.smtp_user, self.smtp_password)
                server.send_message(msg)

        except Exception:
            pass  # Ignora qualquer erro ao enviar o e-mail

    def desconexao_agendada(self):
        """Executa o comando SQL para deletar registros."""
        usuarios_lista = "', '".join(self.usuarios_permitidos)
        comando_sql = f"""
        DELETE FROM R911MOD
        WHERE NUMSEC IN (
            SELECT R911MOD.NUMSEC
            FROM R911MOD
            JOIN R911SEC ON R911MOD.NumSec = R911SEC.NumSec
            WHERE R911SEC.APPUSR NOT IN ('{usuarios_lista}')
        )
        """

        conn = self.conectar_bd()
        if conn:
            try:
                with conn.cursor() as cursor:
                    cursor.execute(comando_sql)
                    conn.commit()
            except oracledb.DatabaseError:
                pass
            finally:
                conn.close()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    janela = AplicativoMonitoramento()
    janela.show()
    sys.exit(app.exec())
